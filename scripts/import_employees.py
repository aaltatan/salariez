import os
from datetime import date
from pathlib import Path
from dataclasses import asdict, dataclass, field, InitVar

from django.core.files import File
from django.utils.text import slugify

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from apps.nationalities.models import Nationality
from apps.areas.models import Area
from apps.departments.models import Department
from apps.job_subtypes.models import JobSubtype
from apps.statuses.models import Status
from apps.positions.models import Position
from apps.employees.models import Employee, Mobile


@dataclass
class EmployeeData:

    position_id: InitVar[str]
    department_id: InitVar[str]
    nationality_id: InitVar[str]
    area_id: InitVar[str]
    job_subtype_id: InitVar[str]
    mobile_number: InitVar[str]

    firstname: str = ''
    father_name: str = ''
    lastname: str = ''
    mother_name: str = ''
    birth_place: str = ''
    birth_date: date | None = None
    national_id: str = ''
    card_id: str = ''
    civil_registry_office: str = ''
    registry_office_name: str = ''
    registry_office_id: str = ''
    gender: str = ''
    current_address: str = ''
    card_date: date | None = None
    religion: str = ''
    hire_date: date | None = None

    salary: int = field(init=False)
    status: str = field(init=False)
    position: str = field(init=False)
    department: str = field(init=False)
    nationality: str = field(init=False)
    area: str = field(init=False)
    job_subtype: str = field(init=False)
    slug: str = field(init=False)
    
    def get_image_name(self, extension: str):
        return self.fullname.strip().replace(' ', '_') + '.' + extension
    
    def __post_init__(
        self, 
        position_id, 
        department_id, 
        nationality_id, 
        area_id, 
        job_subtype_id, 
        mobile_number, 
    ):

        self.nationality = Nationality.objects.get(id=int(nationality_id))
        self.area = Area.objects.get(id=int(area_id))
        self.department = Department.objects.get(id=int(department_id))
        self.job_subtype = JobSubtype.objects.get(id=int(job_subtype_id))
        self.position = Position.objects.get(id=int(position_id))
        self.status = Status.objects.get(id=1)

        self.gender = 'M' if self.gender == 'ذكر' else 'F'
        self.religion = 'M' if self.religion == 'مسلم' else 'C'

        self.salary = 1_000_000
        self.fullname = f'{self.firstname} {self.father_name} {self.lastname}'
        self.slug = slugify(
            f'{self.fullname}-{self.national_id}', 
            allow_unicode=True
        )

        self.mobile = None

        if self.current_address is None:
            self.current_address = ''

        if not self.hire_date:
            self.hire_date = date.today()

        if not self.card_date:
            self.card_date = date.today()

        if mobile_number:
            self.mobile = '0' + str(mobile_number)[3:]
        
        if self.mobile and len(self.mobile) < 10:
            self.mobile = None



def get_employees_data() -> list[EmployeeData]:
    
    file_path = Path('D:').resolve() / 'OneDrive' / 'financial' / 'In_Progress' / 'employees.xlsx'
    
    wb = openpyxl.load_workbook(file_path)
    ws: Worksheet = wb['employees']
    lr = ws.max_row
    table: tuple[tuple[Cell]] = ws[f'A1:V{lr}']

    data = [[cell.value for cell in row] for row in table ]
    
    headers, data = data[0], data[1:]

    employee_data = [
        EmployeeData(**{k:v for k,v in zip(headers, row)}) for row in data
    ]

    return employee_data


def get_onedrive_images() -> dict[str, Path]:
    onedrive_path = Path('D:/onedrive').resolve() / 'financial' / 'PhotosWPU' / 'Profile'

    images_list = os.listdir(onedrive_path)
    images_pathes = [onedrive_path / img for img in images_list]

    images: dict = {}
    for idx, img in enumerate(images_list):
        name = img.split('.')[0]
        images[name] = images_pathes[idx]

    return images


def run() -> None:

    employees = get_employees_data()
    images = get_onedrive_images()

    for emp in employees:

        employee = Employee(**asdict(emp))

        print(employee)

        if images.get(employee.fullname):

            image_path = images[emp.fullname]
            image_file = open(image_path, 'rb')

            image = File(image_file)

            extension = image_path.name.split('.')[-1]

            employee.profile.save(emp.get_image_name(extension), image)

            image_file.close()

        if emp.mobile:
            mobile = Mobile(mobile=emp.mobile, employee=employee)

        employee.save()
        
        if emp.mobile:
            mobile.save()
